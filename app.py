import os
from datetime import datetime
from io import BytesIO

import psycopg2
import psycopg2.extras
from flask import Flask, render_template, request, redirect, send_file, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "troque-esta-chave-secreta")


# =========================
# BANCO POSTGRESQL
# =========================

def get_db():
    database_url = os.environ.get("DATABASE_URL")

    if not database_url:
        raise Exception("DATABASE_URL não configurado no Render.")

    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)

    return psycopg2.connect(
        database_url,
        cursor_factory=psycopg2.extras.RealDictCursor
    )


def criar_banco():
    conn = get_db()
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS usuarios(
            username VARCHAR(100) PRIMARY KEY,
            password TEXT NOT NULL,
            tipo TEXT NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS clientes(
            nome TEXT PRIMARY KEY
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS registros(
            id SERIAL PRIMARY KEY,
            data DATE,
            titulo TEXT,
            cliente TEXT,
            colaborador TEXT,
            nivel TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS historico(
            id SERIAL PRIMARY KEY,
            usuario TEXT,
            acao TEXT,
            data_hora TEXT
        )
    """)

    c.execute("SELECT username FROM usuarios WHERE username=%s", ("Liderseg",))
    admin = c.fetchone()

    if not admin:
        senha_hash = generate_password_hash("123")
        c.execute("""
            INSERT INTO usuarios(username, password, tipo)
            VALUES (%s, %s, %s)
        """, ("Liderseg", senha_hash, "admin"))

    conn.commit()
    conn.close()


def registrar_historico(usuario, acao):
    try:
        conn = get_db()
        c = conn.cursor()

        c.execute("""
            INSERT INTO historico(usuario, acao, data_hora)
            VALUES (%s, %s, %s)
        """, (
            usuario,
            acao,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ))

        conn.commit()
        conn.close()
    except Exception:
        pass


criar_banco()


# =========================
# LOGIN
# =========================

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


class User(UserMixin):
    def __init__(self, username, tipo):
        self.id = username
        self.tipo = tipo


@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM usuarios WHERE username=%s", (user_id,))
    user = c.fetchone()

    conn.close()

    if user:
        return User(user["username"], user["tipo"])

    return None


def is_admin():
    return current_user.is_authenticated and current_user.tipo == "admin"


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db()
        c = conn.cursor()

        c.execute("SELECT * FROM usuarios WHERE username=%s", (username,))
        user = c.fetchone()

        conn.close()

        if user and check_password_hash(user["password"], password):
            login_user(User(user["username"], user["tipo"]))
            registrar_historico(username, "Entrou no sistema")
            return redirect("/")

        flash("Usuário ou senha inválidos.")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    registrar_historico(current_user.id, "Saiu do sistema")
    logout_user()
    return redirect("/login")


# =========================
# DASHBOARD
# =========================

@app.route("/", methods=["GET", "POST"])
@login_required
def index():
    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        if not is_admin():
            conn.close()
            return "Sem permissão"

        c.execute("""
            INSERT INTO registros(data, titulo, cliente, colaborador, nivel)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            request.form["data"],
            request.form["titulo"],
            request.form["cliente"],
            request.form["colaborador"],
            request.form["nivel"]
        ))

        conn.commit()

        registrar_historico(
            current_user.id,
            f"Criou registro para {request.form['colaborador']}"
        )

        conn.close()
        return redirect("/")

    busca = request.args.get("busca", "")
    inicio = request.args.get("inicio", "")
    fim = request.args.get("fim", "")

    sql = "SELECT * FROM registros WHERE 1=1"
    params = []

    if not is_admin():
        sql += " AND colaborador=%s"
        params.append(current_user.id)

    if busca:
        sql += """
            AND (
                titulo ILIKE %s
                OR cliente ILIKE %s
                OR colaborador ILIKE %s
            )
        """
        termo = f"%{busca}%"
        params.extend([termo, termo, termo])

    if inicio:
        sql += " AND data >= %s"
        params.append(inicio)

    if fim:
        sql += " AND data <= %s"
        params.append(fim)

    sql += " ORDER BY id DESC"

    c.execute(sql, params)
    registros = c.fetchall()

    c.execute("SELECT nome FROM clientes ORDER BY nome")
    clientes = c.fetchall()

    c.execute("SELECT username FROM usuarios WHERE tipo='operador' ORDER BY username")
    operadores = c.fetchall()

    # =========================
    # KPIs AVANÇADOS
    # =========================

    filtro_usuario = ""
    filtro_params = []

    if not is_admin():
        filtro_usuario = "WHERE colaborador=%s"
        filtro_params.append(current_user.id)

    c.execute(f"""
        SELECT COUNT(*) AS total
        FROM registros
        {filtro_usuario}
        {"AND" if filtro_usuario else "WHERE"} data = CURRENT_DATE
    """, filtro_params)

    registros_hoje = c.fetchone()["total"] or 0

    c.execute(f"""
        SELECT COUNT(*) AS total
        FROM registros
        {filtro_usuario}
        {"AND" if filtro_usuario else "WHERE"} date_trunc('month', data) = date_trunc('month', CURRENT_DATE)
    """, filtro_params)

    registros_mes = c.fetchone()["total"] or 0

    c.execute(f"""
        SELECT COALESCE(SUM(
            CASE
                WHEN nivel='1' THEN 1.99
                WHEN nivel='2' THEN 2.99
                WHEN nivel='3' THEN 4.99
                WHEN nivel='4' THEN 7.99
                ELSE 0
            END
        ), 0) AS total
        FROM registros
        {filtro_usuario}
    """, filtro_params)

    total_produzido = round(float(c.fetchone()["total"] or 0), 2)

    c.execute(f"""
        SELECT colaborador, COUNT(*) AS total
        FROM registros
        {filtro_usuario}
        GROUP BY colaborador
        ORDER BY total DESC
        LIMIT 1
    """, filtro_params)

    melhor = c.fetchone()
    melhor_operador = melhor["colaborador"] if melhor else "-"

    c.execute(f"""
        SELECT cliente, COUNT(*) AS total
        FROM registros
        {filtro_usuario}
        GROUP BY cliente
        ORDER BY total DESC
        LIMIT 1
    """, filtro_params)

    top_cliente = c.fetchone()
    cliente_top = top_cliente["cliente"] if top_cliente else "-"

    conn.close()

    return render_template(
        "index.html",
        registros=registros,
        clientes=[x["nome"] for x in clientes],
        operadores=[x["username"] for x in operadores],
        busca=busca,
        inicio=inicio,
        fim=fim,
        is_admin=is_admin(),
        registros_hoje=registros_hoje,
        registros_mes=registros_mes,
        total_produzido=f"{total_produzido:.2f}",
        melhor_operador=melhor_operador,
        cliente_top=cliente_top
    )


# =========================
# REGISTROS
# =========================

@app.route("/editar/<int:id>", methods=["GET", "POST"])
@login_required
def editar(id):
    if not is_admin():
        return "Sem permissão"

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM registros WHERE id=%s", (id,))
    registro = c.fetchone()

    if not registro:
        conn.close()
        return "Registro não encontrado"

    if request.method == "POST":
        c.execute("""
            UPDATE registros
            SET data=%s, titulo=%s, cliente=%s, colaborador=%s, nivel=%s
            WHERE id=%s
        """, (
            request.form["data"],
            request.form["titulo"],
            request.form["cliente"],
            request.form["colaborador"],
            request.form["nivel"],
            id
        ))

        conn.commit()
        conn.close()

        registrar_historico(current_user.id, f"Editou registro {id}")
        return redirect("/")

    c.execute("SELECT nome FROM clientes ORDER BY nome")
    clientes = c.fetchall()

    conn.close()

    return render_template(
        "editar.html",
        r=registro,
        clientes=[x["nome"] for x in clientes],
        is_admin=True
    )


@app.route("/excluir/<int:id>")
@login_required
def excluir(id):
    if not is_admin():
        return "Sem permissão"

    conn = get_db()
    c = conn.cursor()

    c.execute("DELETE FROM registros WHERE id=%s", (id,))
    conn.commit()
    conn.close()

    registrar_historico(current_user.id, f"Excluiu registro {id}")

    return redirect("/")


# =========================
# CLIENTES
# =========================

@app.route("/clientes", methods=["GET", "POST"])
@login_required
def clientes():
    if not is_admin():
        return "Acesso negado"

    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        nome = request.form["nome"].strip()

        if nome:
            try:
                c.execute("""
                    INSERT INTO clientes(nome)
                    VALUES (%s)
                    ON CONFLICT (nome) DO NOTHING
                """, (nome,))
                conn.commit()

                registrar_historico(current_user.id, f"Criou cliente {nome}")

            except Exception as e:
                conn.rollback()
                flash(f"Erro ao criar cliente: {e}")

    c.execute("SELECT * FROM clientes ORDER BY nome")
    lista = c.fetchall()

    conn.close()

    return render_template("clientes.html", clientes=lista)


@app.route("/editar_cliente/<nome>", methods=["GET", "POST"])
@login_required
def editar_cliente(nome):
    if not is_admin():
        return "Acesso negado"

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM clientes WHERE nome=%s", (nome,))
    cliente = c.fetchone()

    if not cliente:
        conn.close()
        return redirect("/clientes")

    if request.method == "POST":
        novo_nome = request.form["nome"].strip()

        if novo_nome:
            c.execute(
                "UPDATE clientes SET nome=%s WHERE nome=%s",
                (novo_nome, nome)
            )

            conn.commit()
            conn.close()

            registrar_historico(current_user.id, f"Editou cliente {nome}")
            return redirect("/clientes")

    conn.close()

    return render_template("editar_cliente.html", cliente=cliente)


@app.route("/excluir_cliente/<nome>")
@login_required
def excluir_cliente(nome):
    if not is_admin():
        return "Acesso negado"

    conn = get_db()
    c = conn.cursor()

    c.execute("DELETE FROM clientes WHERE nome=%s", (nome,))
    conn.commit()
    conn.close()

    registrar_historico(current_user.id, f"Excluiu cliente {nome}")

    return redirect("/clientes")


# =========================
# USUÁRIOS
# =========================

@app.route("/usuarios", methods=["GET", "POST"])
@login_required
def usuarios():
    if not is_admin():
        return "Acesso negado"

    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        tipo = request.form["tipo"]

        if username and password:
            try:
                senha_hash = generate_password_hash(password)

                c.execute("""
                    INSERT INTO usuarios(username, password, tipo)
                    VALUES (%s, %s, %s)
                """, (username, senha_hash, tipo))

                conn.commit()

                registrar_historico(current_user.id, f"Criou usuário {username}")

            except Exception as e:
                conn.rollback()
                flash(f"Erro ao criar usuário: {e}")

    c.execute("SELECT username, tipo FROM usuarios ORDER BY username")
    lista = c.fetchall()

    conn.close()

    return render_template("usuarios.html", usuarios=lista)


@app.route("/editar_usuario/<username>", methods=["GET", "POST"])
@login_required
def editar_usuario(username):
    if not is_admin():
        return "Acesso negado"

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM usuarios WHERE username=%s", (username,))
    usuario = c.fetchone()

    if not usuario:
        conn.close()
        return redirect("/usuarios")

    if request.method == "POST":
        nova_senha = request.form["password"]
        tipo = request.form["tipo"]

        if nova_senha:
            senha_hash = generate_password_hash(nova_senha)

            c.execute("""
                UPDATE usuarios
                SET password=%s, tipo=%s
                WHERE username=%s
            """, (senha_hash, tipo, username))

        else:
            c.execute("""
                UPDATE usuarios
                SET tipo=%s
                WHERE username=%s
            """, (tipo, username))

        conn.commit()
        conn.close()

        registrar_historico(current_user.id, f"Editou usuário {username}")

        return redirect("/usuarios")

    conn.close()

    return render_template("editar_usuario.html", usuario=usuario)


@app.route("/alterar_senha/<username>", methods=["POST"])
@login_required
def alterar_senha(username):
    if not is_admin():
        return "Sem permissão"

    nova_senha = request.form["nova_senha"]

    if not nova_senha:
        return redirect("/usuarios")

    senha_hash = generate_password_hash(nova_senha)

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        UPDATE usuarios
        SET password=%s
        WHERE username=%s
    """, (senha_hash, username))

    conn.commit()
    conn.close()

    registrar_historico(current_user.id, f"Alterou senha do usuário {username}")

    return redirect("/usuarios")


@app.route("/excluir_usuario/<username>")
@login_required
def excluir_usuario(username):
    if not is_admin():
        return "Acesso negado"

    if username == current_user.id:
        return "Não é permitido excluir seu próprio usuário."

    conn = get_db()
    c = conn.cursor()

    c.execute("DELETE FROM usuarios WHERE username=%s", (username,))
    conn.commit()
    conn.close()

    registrar_historico(current_user.id, f"Excluiu usuário {username}")

    return redirect("/usuarios")


# =========================
# HISTÓRICO
# =========================

@app.route("/historico")
@login_required
def historico():
    if not is_admin():
        return "Acesso negado"

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM historico ORDER BY id DESC")
    dados = c.fetchall()

    conn.close()

    return render_template("historico.html", historico=dados)


@app.route("/limpar_historico")
@login_required
def limpar_historico():
    if not is_admin():
        return "Acesso negado"

    conn = get_db()
    c = conn.cursor()

    c.execute("DELETE FROM historico")
    conn.commit()
    conn.close()

    registrar_historico(current_user.id, "Limpou o histórico")

    return redirect("/historico")


# =========================
# RANKING
# =========================

@app.route("/ranking")
@login_required
def ranking():
    if not is_admin():
        return "Acesso negado"

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT 
            colaborador,
            COUNT(*) AS total,
            COALESCE(SUM(
                CASE
                    WHEN nivel='1' THEN 1.99
                    WHEN nivel='2' THEN 2.99
                    WHEN nivel='3' THEN 4.99
                    WHEN nivel='4' THEN 7.99
                    ELSE 0
                END
            ), 0) AS valor_total
        FROM registros
        WHERE colaborador IS NOT NULL
        GROUP BY colaborador
        ORDER BY total DESC
    """)

    dados = c.fetchall()

    conn.close()

    return render_template(
        "ranking.html",
        ranking=dados
    )


# =========================
# GRÁFICO
# =========================

@app.route("/grafico")
@login_required
def grafico():
    conn = get_db()
    c = conn.cursor()

    operador = request.args.get("operador", "todos")

    if is_admin():
        if operador == "todos":
            c.execute("SELECT * FROM registros")
        else:
            c.execute("SELECT * FROM registros WHERE colaborador=%s", (operador,))
    else:
        operador = current_user.id
        c.execute("SELECT * FROM registros WHERE colaborador=%s", (current_user.id,))

    dados = c.fetchall()

    n1 = len([x for x in dados if x["nivel"] == "1"])
    n2 = len([x for x in dados if x["nivel"] == "2"])
    n3 = len([x for x in dados if x["nivel"] == "3"])
    n4 = len([x for x in dados if x["nivel"] == "4"])

    v1 = round(n1 * 1.99, 2)
    v2 = round(n2 * 2.99, 2)
    v3 = round(n3 * 4.99, 2)
    v4 = round(n4 * 7.99, 2)

    total = round(v1 + v2 + v3 + v4, 2)

    operadores = []

    if is_admin():
        c.execute("""
            SELECT DISTINCT colaborador
            FROM registros
            WHERE colaborador IS NOT NULL
            ORDER BY colaborador
        """)
        operadores = [x["colaborador"] for x in c.fetchall()]

    conn.close()

    return render_template(
        "grafico.html",
        n1=n1,
        n2=n2,
        n3=n3,
        n4=n4,
        v1=v1,
        v2=v2,
        v3=v3,
        v4=v4,
        total=total,
        operadores=operadores,
        operador_selecionado=operador,
        is_admin=is_admin()
    )


# =========================
# EXPORTAR EXCEL
# =========================

@app.route("/exportar_excel")
@login_required
def exportar_excel():
    conn = get_db()
    c = conn.cursor()

    if is_admin():
        c.execute("SELECT * FROM registros ORDER BY id DESC")
    else:
        c.execute(
            "SELECT * FROM registros WHERE colaborador=%s ORDER BY id DESC",
            (current_user.id,)
        )

    registros = c.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    ws.append([
        "ID",
        "Data",
        "Título",
        "Cliente",
        "Operador",
        "Nível"
    ])

    for r in registros:
        ws.append([
            r["id"],
            str(r["data"]) if r["data"] else "",
            r["titulo"],
            r["cliente"],
            r["colaborador"],
            r["nivel"]
        ])

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name="registros.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================
# IMPORTAR REGISTROS POR EXCEL
# =========================

@app.route("/importar_excel", methods=["GET", "POST"])
@login_required
def importar_excel():
    if not is_admin():
        return "Acesso negado"

    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        data_lancamento = request.form.get("data")

        if not arquivo or arquivo.filename == "":
            flash("Selecione um arquivo Excel.")
            return redirect("/importar_excel")

        if not data_lancamento:
            flash("Informe a data dos registros.")
            return redirect("/importar_excel")

        try:
            wb = load_workbook(arquivo)
            ws = wb.active

            headers = {}
            linha_cabecalho = None

            for row in range(1, 10):
                for col in range(1, ws.max_column + 1):
                    valor = ws.cell(row=row, column=col).value

                    if valor:
                        nome = str(valor).strip().lower()

                        if "título" in nome or "titulo" in nome:
                            headers["titulo"] = col
                            linha_cabecalho = row

                        elif "colaborador" in nome:
                            headers["colaborador"] = col

                        elif "pontua" in nome:
                            headers["nivel"] = col

                        elif "posto" in nome:
                            headers["cliente"] = col

                if all(k in headers for k in ["titulo", "colaborador", "nivel", "cliente"]):
                    break

            if not all(k in headers for k in ["titulo", "colaborador", "nivel", "cliente"]):
                flash("A planilha precisa ter as colunas: Título, Colaborador, Pontuação e Posto Trabal.")
                return redirect("/importar_excel")

            conn = get_db()
            c = conn.cursor()

            importados = 0

            for row in range(linha_cabecalho + 1, ws.max_row + 1):
                titulo = ws.cell(row=row, column=headers["titulo"]).value
                colaborador = ws.cell(row=row, column=headers["colaborador"]).value
                nivel = ws.cell(row=row, column=headers["nivel"]).value
                cliente = ws.cell(row=row, column=headers["cliente"]).value

                if not titulo or not colaborador or not nivel or not cliente:
                    continue

                titulo = str(titulo).strip()
                colaborador = str(colaborador).strip()
                nivel = str(nivel).strip()
                cliente = str(cliente).strip()

                c.execute("""
                    INSERT INTO clientes(nome)
                    VALUES (%s)
                    ON CONFLICT (nome) DO NOTHING
                """, (cliente,))

                c.execute("""
                    INSERT INTO registros(data, titulo, cliente, colaborador, nivel)
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    data_lancamento,
                    titulo,
                    cliente,
                    colaborador,
                    nivel
                ))

                importados += 1

            conn.commit()
            conn.close()

            registrar_historico(
                current_user.id,
                f"Importou {importados} registros via Excel"
            )

            flash(f"{importados} registros importados com sucesso.")
            return redirect("/")

        except Exception as e:
            flash(f"Erro ao importar Excel: {e}")
            return redirect("/importar_excel")

    return render_template("importar_excel.html")
# =========================
# DIAGNÓSTICO
# =========================

@app.route("/diagnostico")
@login_required
def diagnostico():
    try:
        conn = get_db()
        c = conn.cursor()

        c.execute("SELECT COUNT(*) AS total FROM clientes")
        total_clientes = c.fetchone()["total"]

        c.execute("SELECT COUNT(*) AS total FROM usuarios")
        total_usuarios = c.fetchone()["total"]

        c.execute("SELECT COUNT(*) AS total FROM registros")
        total_registros = c.fetchone()["total"]

        c.execute("SELECT current_database() AS banco")
        banco = c.fetchone()["banco"]

        conn.close()

        return f"""
        <h2>Diagnóstico do Banco</h2>
        <b>Banco:</b> {banco}<br><br>
        <b>Total de Clientes:</b> {total_clientes}<br>
        <b>Total de Usuários:</b> {total_usuarios}<br>
        <b>Total de Registros:</b> {total_registros}<br><br>
        <b>DATABASE_URL configurada:</b>
        {'SIM' if os.environ.get('DATABASE_URL') else 'NÃO'}
        """

    except Exception as e:
        return f"ERRO: {e}"


# =========================
# EXECUTAR
# =========================

if __name__ == "__main__":
    app.run(debug=True)